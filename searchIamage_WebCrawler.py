# -*- coding: utf-8 -*-
"""
Created on Fri Nov 20 19:31:08 2023

@author: avend
"""

# -*- coding: utf-8 -*-
import requests
import tkinter as tk
from tkinter import filedialog, messagebox, Label, Button, Scrollbar, Entry, Canvas, Frame, IntVar, Scale
from bs4 import BeautifulSoup
from PIL import Image, ImageTk, UnidentifiedImageError
import io
import os
import openpyxl

class ImageDownloader:
    def __init__(self, master):
        self.master = master
        master.title("Buscador de imágenes")

        # Configurar tamaño mínimo para la ventana
        master.minsize(width=850, height=500)
        master.maxsize(width=850, height=650)

        # Configuración de la interfaz de usuario
        self.setup_ui()

        # Lista para almacenar imágenes y sus datos
        self.images = []
        self.image_data_list = []

    def setup_ui(self):        
        # Etiqueta y entrada para la palabra de búsqueda
        self.label = Label(self.master, text="Ingresa la imagen a buscar:", bg="#aedbf2")
        self.label.pack(padx=10, pady=10)

        self.entry = Entry(self.master, bg="#ffffff")
        self.entry.pack(padx=10, pady=10)

        # Botón para buscar imágenes
        self.button_search = Button(self.master, text="Buscar imágenes", height=1, width=15, command=self.load_images, bg="#5cb85c", fg="#ffffff")
        self.button_search.pack(side=tk.TOP, pady=(10, 0))

        # Canvas para mostrar las imágenes descargadas con la barra de desplazamiento
        self.canvas = Canvas(self.master, bg="#f0f0f0")
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Configurar la barra de desplazamiento vertical
        self.scrollbar = Scrollbar(self.master, orient="vertical", command=self.canvas.yview)
        self.scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.canvas.configure(yscrollcommand=self.scrollbar.set)

        # Marco para contener las miniaturas de imágenes en el canvas
        self.images_frame = Frame(self.canvas, bg="#ffffff")
        self.canvas.create_window((0, 0), window=self.images_frame, anchor=tk.NW)

        # Etiqueta y control deslizante para el número de imágenes
        self.image_count_label = Label(self.master, text="Número de imágenes:", bg="#aedbf2")
        self.image_count_label.pack(pady=(100, 5))

        self.image_count_var = IntVar(value=20)  # Valor predeterminado
        self.image_count_slider = Scale(self.master, from_=1, to=20, orient=tk.HORIZONTAL, variable=self.image_count_var, length=150, bg="#f0f0f0", troughcolor="#aedbf2")
        self.image_count_slider.pack(pady=(0, 10))
    
        # Botón para descargar imágenes seleccionadas
        self.download_button = Button(self.master, text="Descargar imágenes seleccionadas", command=self.download_selected_images, bg="#5cb85c", fg="#ffffff")
        self.download_button.pack(padx=10, pady=25)
        self.download_button.config(state=tk.DISABLED)  # Deshabilitar el botón al inicio

    def load_images(self):
        # Método para cargar las miniaturas de las imágenes sin descargarlas
        self.download_button.config(state=tk.DISABLED)  # Deshabilitar el botón al inicio
        self.clear_images()  # Llamar a clear_images también al inicio de load_images
        self.image_data_list = []

        # Obtener el número de imágenes deseadas del slider
        num_images = self.image_count_var.get()

        # Obtener la palabra clave para la búsqueda
        self.keywords = self.entry.get()
        url = f"https://www.google.com/search?q={self.keywords}&source=lnms&tbm=isch"
        
        # # Búsqueda en DuckDuckGo
        # duckduckgo_keywords = f"{self.keywords} más específico"
        # duckduckgo_url = f"https://duckduckgo.com/?q={duckduckgo_keywords}&iax=images&ia=images"
        # duckduckgo_results = self.search_images(duckduckgo_url, num_images)
        
        # # Búsqueda en Unsplash
        # unsplash_keywords = f"{self.keywords} bonitas"
        # unsplash_url = f"https://unsplash.com/s/photos/{unsplash_keywords}"
        # unsplash_results = self.search_images(unsplash_url, num_images)

        try:
            response = requests.get(url)
            response.raise_for_status()
            soup = BeautifulSoup(response.content, 'html.parser')
            img_tags = soup.find_all('img')

            # Nueva estructura para organizar las imágenes en filas de tres
            row_num = 0
            col_num = 0
            self.images = []

            for i, img in enumerate(img_tags):
                img_url = img['src']
                if img_url.startswith('https') and i < num_images + 1:
                    try:
                        # Descargar la imagen y almacenar sus datos
                        response = requests.get(img_url)
                        response.raise_for_status()
                        img_data = response.content
                        img = Image.open(io.BytesIO(img_data))
                        # Redimensionar la imagen a 150x150 píxeles
                        img.thumbnail((150, 150))

                        # Almacenar la imagen y sus datos en la lista
                        self.image_data_list.append(img_data)

                        # Miniatura de la imagen
                        img = ImageTk.PhotoImage(img)
                        thumbnail_label = Label(self.images_frame, image=img, bg="#ffffff")
                        thumbnail_label.image = img
                        thumbnail_label.grid(row=row_num, column=col_num, padx=20, pady=15)

                        # Checkbox para cada imagen
                        checkbox_var = tk.BooleanVar()
                        checkbox = tk.Checkbutton(self.images_frame, variable=checkbox_var, bg="#ffffff")

                        # Asegurarse de almacenar la variable en la tupla de imágenes
                        self.images.append((thumbnail_label, checkbox_var))

                        # Ajustar el valor de pady para reducir el espacio vertical
                        checkbox.grid(row=row_num + 1, column=col_num, pady=(0, 1))

                        # Incrementar el número de columna
                        col_num += 1

                        # Si se alcanza el límite de tres columnas, pasar a la siguiente fila
                        if col_num == 3:
                            col_num = 0
                            row_num += 2  

                    except (UnidentifiedImageError, requests.RequestException) as e:
                        print("Error al obtener más imágenes:", str(e))

            # Configurar el área desplazable del lienzo después de una configuración
            self.canvas.configure(scrollregion=self.canvas.bbox("all"))

            # Configurar el área desplazable del lienzo después de una configuración
            self.update_scrollregion()
            self.download_button.config(state=tk.NORMAL)

        except Exception as e:
            print(f"Error al obtener más imágenes: {e}")

    def clear_images(self):
        # Método para destruir los widgets de las imágenes anteriores
        for widget in self.images_frame.winfo_children():
            widget.destroy()

        # Limpiar la lista de imágenes
        self.images = []

    def update_scrollregion(self):
        # Método para actualizar la barra deslizadora
        self.canvas.update_idletasks()
        self.canvas.config(scrollregion=self.canvas.bbox("all"))

    def download_selected_images(self):
        selected_images = [image for image, (thumbnail_label, checkbox_var) in zip(self.images, self.images) if checkbox_var.get()]

        if not selected_images:
            messagebox.showinfo("Ninguna imagen seleccionada", "Por favor, seleccione al menos una imagen para descargar.")
            return

        # Crear una carpeta para las imágenes seleccionadas
        self.folder_name = filedialog.askdirectory()
        selected_folder = os.path.join(self.folder_name, f"{self.keywords}_Images")
        os.makedirs(selected_folder, exist_ok=True)
        
        # Crear un libro de Excel y seleccionar la primera hoja
        wb = openpyxl.Workbook()
        sheet = wb.active

        # Añadir encabezados a las columnas
        sheet["A1"] = "Imagen"
        sheet["B1"] = "URL de Búsqueda"
        
        for count, (thumbnail_label, _) in enumerate(selected_images):
            # Obtener el dato de la imagen original desde la lista
            img_data = self.image_data_list[count]

            # Obtener el formato de la imagen original (si es posible)
            img_format = "png"  # Cambia esto según tus necesidades

            # Guardar la imagen seleccionada en la nueva carpeta con el formato correcto
            img_path = os.path.join(selected_folder, f"image_{count + 1}.{img_format}")
            with open(img_path, 'wb') as img_file:
                img_file.write(img_data)
            
            # Añadir la información al libro de Excel
            sheet.cell(row=count + 2, column=1, value=f"image_{count + 1}")
            sheet.cell(row=count + 2, column=2, value=f"https://www.google.com/search?q={self.keywords}&tbm=isch")
            
        # Guardar el libro de Excel
        excel_file_path = os.path.join(self.folder_name, f"{self.keywords}_Search_URLs.xlsx")
        wb.save(excel_file_path)

        # Mostrar mensaje de descarga completa
        messagebox.showinfo("Descarga completa", f"Las imágenes seleccionadas han sido descargadas en la carpeta {self.keywords}_Images.")

if __name__ == "__main__":
    root = tk.Tk()
    app = ImageDownloader(root)
    root.mainloop()
